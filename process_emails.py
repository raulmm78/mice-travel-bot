from __future__ import annotations

import csv
import imaplib
import html
import json
import os
import random
import re
import socket
import smtplib
import subprocess
import sys
import threading
import time
import traceback
import unicodedata
import urllib.error
import urllib.request
from copy import copy
from dataclasses import dataclass, asdict
from datetime import date, datetime
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
EMAIL_DIR = BASE_DIR / "emails"
ASSETS_DIR = BASE_DIR / "assets"
ENV_PATH = BASE_DIR / ".env"
OUTPUT_DIR = BASE_DIR.parents[1] / "outputs"
XLSX_PATH = OUTPUT_DIR / "viajes_demo.xlsx"
CSV_PATH = OUTPUT_DIR / "viajes_demo.csv"
JSON_PATH = OUTPUT_DIR / "viajes_demo.json"
LOG_PATH = OUTPUT_DIR / "email_agent.log"
EVENT_OUTPUT_DIR = OUTPUT_DIR / "eventos"
PROCESSED_IDS_FILENAME = "_bot_processed_message_ids.json"
DEFAULT_NN_TEMPLATE_PATH = Path("/Users/raulmartinez/Library/Containers/com.apple.mail/Data/Library/Mail Downloads/084DBDBC-2ECD-4DB2-BE23-DF294F19A3AB/LISTADO PARA VOLCAR LOS DATOS NN.xlsx")
WATCH_INTERVAL_SECONDS = int(os.getenv("WATCH_INTERVAL_SECONDS", "300"))
SERVER_HOST = "127.0.0.1"
DEFAULT_SERVER_PORT = 8765
OPENAI_API_URL = "https://api.openai.com/v1/responses"
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4.1-mini")
OPENAI_TIMEOUT_SECONDS = int(os.getenv("OPENAI_TIMEOUT_SECONDS", "20"))
MANDATORY_FIELDS = ("nombre", "dni", "origen", "destino", "fecha_viaje")
BOT_ADDED_FONT_COLOR = "0070C0"
EXCEL_DATE_FORMAT = "dd/mm/yy"
EXCEL_MONTH_FORMAT = "mm/yyyy"
BOT_LOCK = threading.Lock()
BOT_STOP_EVENT = threading.Event()
BOT_THREAD: threading.Thread | None = None
BOT_ACTIVE = False
OUTPUT_FIELDS = [
    "email_id",
    "source_message_id",
    "nombre",
    "apellidos",
    "dni",
    "email",
    "telefono",
    "evento",
    "origen",
    "destino",
    "fecha_viaje",
    "hora_preferida",
    "preferencia",
    "observaciones",
    "estado",
    "dudas",
    "delegado_rma",
    "gerente",
    "cost_center",
    "internal_order",
    "selas_id",
    "hospital",
    "especialidad",
    "residente",
    "socio",
    "ponente_oyente",
    "consentimiento_firmado",
    "curso",
    "congreso",
    "departamento",
    "cost_center_completo",
    "proyecto",
    "restricciones_alimentarias",
    "descripcion_servicio",
    "conex_ida",
    "desplazamientos_ida",
    "desplazamientos_vuelta",
    "conex_regreso",
    "hotel",
    "hotel_in",
    "hotel_out",
    "inscripcion",
]
DEMO_FIRST_NAMES = ["Lucia", "Pablo", "Elena", "Sergio", "Carmen", "Diego", "Irene", "Raul", "Nuria", "Alvaro"]
DEMO_LAST_NAMES = [
    "Sanchez Martin",
    "Herrera Lopez",
    "Navarro Ruiz",
    "Torres Garcia",
    "Vidal Moreno",
    "Castro Perez",
    "Romero Diaz",
    "Molina Serra",
    "Ortega Leon",
    "Ramos Gil",
]
DEMO_CITIES = ["Madrid", "Barcelona", "Valencia", "Sevilla", "Malaga", "Bilbao", "Zaragoza", "Alicante", "Vigo", "Granada"]
DEMO_EVENTS = ["cardiologia", "oncologia", "dermatologia", "neurologia", "farmacia hospitalaria"]
DEMO_CONGRESSES = ["Congreso IMS", "Congreso ESC", "Congreso EASD", "Congreso SEN", "Congreso SEGO"]
DEMO_HOSPITALS = ["Hospital Universitario La Paz", "Hospital Clinic", "Hospital Virgen del Rocio", "Menoclinica by Palacios", "Hospital Quiron"]


@dataclass
class TravelRequest:
    email_id: str
    source_message_id: str = ""
    nombre: str = ""
    apellidos: str = ""
    dni: str = ""
    email: str = ""
    telefono: str = ""
    evento: str = ""
    origen: str = ""
    destino: str = ""
    fecha_viaje: str = ""
    hora_preferida: str = ""
    preferencia: str = ""
    observaciones: str = ""
    estado: str = "pendiente_revision"
    dudas: str = ""
    delegado_rma: str = ""
    gerente: str = ""
    cost_center: str = ""
    internal_order: str = ""
    selas_id: str = ""
    hospital: str = ""
    especialidad: str = ""
    residente: str = ""
    socio: str = ""
    ponente_oyente: str = ""
    consentimiento_firmado: str = ""
    curso: str = ""
    congreso: str = ""
    departamento: str = ""
    cost_center_completo: str = ""
    proyecto: str = ""
    restricciones_alimentarias: str = ""
    descripcion_servicio: str = ""
    conex_ida: str = ""
    desplazamientos_ida: str = ""
    desplazamientos_vuelta: str = ""
    conex_regreso: str = ""
    hotel: str = ""
    hotel_in: str = ""
    hotel_out: str = ""
    inscripcion: str = ""


def load_env_file(path: Path = ENV_PATH) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        clean = line.strip()
        if not clean or clean.startswith("#") or "=" not in clean:
            continue
        key, value = clean.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


EXTRACTION_PROMPT = """
Eres un asistente experto en operaciones de viajes para eventos corporativos en España.
Tu tarea es leer un email de un asistente y extraer datos estructurados para actualizar una hoja de cálculo.

Reglas importantes:
- Devuelve solo datos presentes o inferidos de forma muy segura a partir del email.
- No inventes información. Si un dato no aparece o es ambiguo, usa cadena vacía en ese campo.
- Extrae nombres y apellidos por separado cuando sea posible.
- Normaliza fecha_viaje a formato ISO YYYY-MM-DD. Si falta el año y el email no permite deducirlo con seguridad, deja el campo vacío.
- Mantén hora_preferida como texto si expresa un rango o preferencia aproximada, por ejemplo "07:00-09:00" o "mañana".
- En preferencia incluye preferencias de transporte, asiento, tren directo, vuelo, horarios, etc.
- En observaciones incluye cualquier detalle operativo que no encaje en otros campos.
- En dudas explica brevemente ambigüedades o información faltante detectada.
- El campo estado debe ser "ok" solo si están completos nombre, dni, origen, destino y fecha_viaje; en caso contrario debe ser "pendiente_revision".

Campos de salida:
nombre, apellidos, dni, email, telefono, evento, origen, destino, fecha_viaje,
hora_preferida, preferencia, observaciones, estado, dudas.
""".strip()


OPENAI_JSON_SCHEMA = {
    "name": "travel_email_extraction",
    "strict": True,
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "nombre": {"type": "string"},
            "apellidos": {"type": "string"},
            "dni": {"type": "string"},
            "email": {"type": "string"},
            "telefono": {"type": "string"},
            "evento": {"type": "string"},
            "origen": {"type": "string"},
            "destino": {"type": "string"},
            "fecha_viaje": {"type": "string"},
            "hora_preferida": {"type": "string"},
            "preferencia": {"type": "string"},
            "observaciones": {"type": "string"},
            "estado": {"type": "string", "enum": ["ok", "pendiente_revision"]},
            "dudas": {"type": "string"},
        },
        "required": [
            "nombre",
            "apellidos",
            "dni",
            "email",
            "telefono",
            "evento",
            "origen",
            "destino",
            "fecha_viaje",
            "hora_preferida",
            "preferencia",
            "observaciones",
            "estado",
            "dudas",
        ],
    },
}


def read_email(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def first_match(pattern: str, text: str, flags: int = re.IGNORECASE) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else ""


def clean_value(raw: str) -> str:
    return re.sub(r"\s+", " ", raw.replace("**", "").strip())


def label_match(label: str, text: str) -> str:
    pattern = rf"^[ \t]*{re.escape(label)}[ \t]*:[ \t]*(.*?)[ \t]*$"
    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    return clean_value(match.group(1)) if match else ""


def normalize_date(raw: str, default_year: str | None = "2026") -> str:
    raw = raw.strip().lower()
    if not raw:
        return ""

    month_names = {
        "enero": "01",
        "febrero": "02",
        "marzo": "03",
        "abril": "04",
        "mayo": "05",
        "junio": "06",
        "julio": "07",
        "agosto": "08",
        "septiembre": "09",
        "setiembre": "09",
        "octubre": "10",
        "noviembre": "11",
        "diciembre": "12",
    }

    numeric = re.search(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", raw)
    if numeric:
        day, month, year = numeric.groups()
        if not year and not default_year:
            return raw
        year = year or default_year
        if len(year) == 2:
            year = "20" + year
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    named = re.search(r"(\d{1,2})\s+de\s+([a-záéíóúñ]+)", raw)
    if named and named.group(2) in month_names:
        if not default_year:
            return raw
        return f"{default_year}-{month_names[named.group(2)]}-{int(named.group(1)):02d}"

    return raw


def normalize_dni(raw: str) -> str:
    match = re.search(r"\b([0-9]{7,8})\s*[- ]?\s*([A-Z])\b", raw.upper())
    return f"{match.group(1)}{match.group(2)}" if match else raw.strip().upper()


def normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", raw)
    return digits if len(digits) >= 9 else raw.strip()


MONTH_WORDS = "enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre"


def service_date(day: str, month: str, fallback_year: str = "2026") -> str:
    return normalize_date(f"{day} de {month}", default_year=fallback_year)


def compact_service_text(raw: str) -> str:
    text = raw.replace("VueloVUELTA", "Vuelo VUELTA")
    return re.sub(r"\s+", " ", text).strip()


def stop_at_next_service_block() -> str:
    return (
        r"(?=(?:\.?\s+-?\s*)?(?:Vuelo\s+directo|Vuelo\s+VUELTA|Vuelo\s+IDA|Alojamiento|"
        r"Tren\b|Traslado\b|Turista\b|Preferencia\b|Por favor\b|No es necesario\b|"
        r"No precisa\b|Inscripci[oó]n\b|$))"
    )


def clean_service_fragment(raw: str) -> str:
    return clean_value(raw.strip(" .-"))


def first_upper(value: str) -> str:
    value = str(value or "").strip()
    return value[:1].upper() + value[1:] if value else ""


def title_case_display(value: str) -> str:
    value = clean_value(str(value or ""))
    acronyms = {"IMS", "ESC", "EASD", "SEN", "SEGO", "DNI", "NIF", "AVIOS"}

    def format_word(match: re.Match[str]) -> str:
        word = match.group(0)
        if word.upper() in acronyms:
            return word.upper()
        return word[:1].upper() + word[1:].lower()

    return re.sub(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+", format_word, value)


def sentence_case_display(value: str) -> str:
    value = clean_value(str(value or "").strip(" -"))
    if not value:
        return ""
    return value[:1].upper() + value[1:]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in normalized if not unicodedata.combining(char))


def excel_text(value: str, *, email: bool = False) -> str:
    value = clean_value(str(value or ""))
    if email:
        return strip_accents(value).lower()
    return strip_accents(value).upper()


def parse_iso_date(value: str) -> date | str:
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return value


def billing_month(value: str) -> date | str:
    parsed = parse_iso_date(value)
    if isinstance(parsed, date):
        return date(parsed.year, parsed.month, 1)
    return ""


def safe_filename(value: str) -> str:
    value = strip_accents(value).upper()
    value = re.sub(r"[^A-Z0-9]+", "_", value).strip("_")
    return value[:80] or "SIN_EVENTO"


def clean_message_id(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        return ""
    return value.strip("<>").strip().lower()


def filename_text(value: str) -> str:
    value = strip_accents(value).upper()
    value = re.sub(r"[^A-Z0-9 ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:90] or "SIN EVENTO"


def filename_date(value: str) -> str:
    parsed = parse_iso_date(value)
    if isinstance(parsed, date):
        return parsed.strftime("%d-%m-%y")
    return "SIN FECHA"


def event_workbook_filename(event_name: str, rows: list[TravelRequest]) -> str:
    dates = sorted(
        row.fecha_viaje
        for row in rows
        if isinstance(parse_iso_date(row.fecha_viaje), date)
    )
    first_date = dates[0] if dates else ""
    return f"NO ENVIAR ---- {filename_text(event_name)} {filename_date(first_date)}.xlsx"


def manager_from_cc(text: str) -> str:
    cc_line = first_match(r"^(?:CC|Cc|Copia):\s*(.+)$", text, re.IGNORECASE | re.MULTILINE)
    if not cc_line:
        return ""
    initials = []
    for email_address in re.findall(r"([A-Z0-9._%+-]+)@(?:novonordisk|novonordisk|novonordisk)\.com", cc_line, re.IGNORECASE):
        local = email_address.split("@")[0].strip()
        token = re.sub(r"[^A-Z]", "", local.upper())
        if 2 <= len(token) <= 6:
            initials.append(token)
    if not initials:
        for email_address in re.findall(r"([A-Z0-9._%+-]+)@[A-Z0-9.-]+", cc_line, re.IGNORECASE):
            token = re.sub(r"[^A-Z]", "", email_address.split("@")[0].upper())
            if 2 <= len(token) <= 6:
                initials.append(token)
    return " / ".join(dict.fromkeys(initials))


def extract_service_details(request: TravelRequest) -> None:
    text = request.descripcion_servicio
    if not text:
        return

    year = request.fecha_viaje[:4] if re.fullmatch(r"\d{4}-\d{2}-\d{2}", request.fecha_viaje) else "2026"
    normalized = compact_service_text(text)
    service_stop = stop_at_next_service_block()

    hotel = re.search(
        rf"Alojamiento\s+del\s+(\d{{1,2}})(?:\s+de\s+({MONTH_WORDS}))?\s+al\s+(\d{{1,2}})\s+de\s+({MONTH_WORDS})(?:\s+en\s+(.+?))?{service_stop}",
        normalized,
        re.IGNORECASE,
    )
    if hotel:
        start_month = hotel.group(2) or hotel.group(4)
        request.hotel_in = service_date(hotel.group(1), start_month, year)
        request.hotel_out = service_date(hotel.group(3), hotel.group(4), year)
        hotel_text = clean_service_fragment(hotel.group(5) or "")
        request.hotel = hotel_text or "Hotel seleccionado"
    else:
        one_night = re.search(
            rf"Alojamiento\s+una\s+noche\s+en\s+(.+?),\s+entrada\s+(\d{{1,2}})\s+de\s+({MONTH_WORDS})\s+y\s+salida\s+(\d{{1,2}})\s+de\s+({MONTH_WORDS})",
            normalized,
            re.IGNORECASE,
        )
        if one_night:
            request.hotel = f"Alojamiento en {clean_service_fragment(one_night.group(1))}"
            request.hotel_in = service_date(one_night.group(2), one_night.group(3), year)
            request.hotel_out = service_date(one_night.group(4), one_night.group(5), year)

    ida = re.search(
        rf"Vuelo\s+IDA\s+(.+?){service_stop}",
        normalized,
        re.IGNORECASE,
    )
    if ida:
        request.desplazamientos_ida = clean_service_fragment(ida.group(1))

    vuelta = re.search(
        rf"Vuelo\s+VUELTA\s+(.+?){service_stop}",
        normalized,
        re.IGNORECASE,
    )
    if vuelta:
        request.desplazamientos_vuelta = clean_service_fragment(vuelta.group(1))

    trains = [
        clean_service_fragment(match.group(1))
        for match in re.finditer(rf"Tren\s+(.+?){service_stop}", normalized, re.IGNORECASE)
    ]
    if trains and not request.desplazamientos_ida:
        request.desplazamientos_ida = trains[0]
    if len(trains) > 1 and not request.desplazamientos_vuelta:
        request.desplazamientos_vuelta = trains[1]

    if re.search(r"\bvuelo\s+directo\b", normalized, re.IGNORECASE):
        request.conex_ida = "Directo"
        request.conex_regreso = "Directo"
    elif re.search(r"\b(con\s+escala|escala)\b", normalized, re.IGNORECASE):
        request.conex_ida = "Con escala"
        request.conex_regreso = "Con escala"

    if re.search(r"no\s+es\s+necesario\s+tramitar\s+inscripci[oó]n", normalized, re.IGNORECASE):
        request.inscripcion = "No tramitar"
    elif re.search(r"\binscripci[oó]n\b", normalized, re.IGNORECASE):
        request.inscripcion = "Revisar"

    if not request.destino:
        route = re.search(r"\b([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+)\s*-\s*([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+)\b", normalized)
        if route:
            request.destino = route.group(2)


def extract_date(text: str) -> str:
    candidates = [
        first_match(r"(?:dia|día|sabado|sábado)\s+([0-9]{1,2}[/-][0-9]{1,2}(?:[/-][0-9]{2,4})?|\d{1,2}\s+de\s+[a-záéíóúñ]+)", text),
        first_match(r"\b(\d{1,2}\s+de\s+[a-záéíóúñ]+)\b", text),
        first_match(r"\bel\s+([0-9]{1,2}[/-][0-9]{1,2}(?:[/-][0-9]{2,4})?)\b", text),
        first_match(r"\b([0-9]{1,2}[/-][0-9]{1,2}/[0-9]{2,4})\b", text),
    ]
    for candidate in candidates:
        normalized = normalize_date(candidate)
        if normalized:
            return normalized
    return ""


def guess_name(text: str) -> tuple[str, str]:
    explicit = first_match(r"soy\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][\wáéíóúñ]+){1,3})", text, re.IGNORECASE)
    if not explicit:
        explicit = first_match(r"\n\s*([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][\wáéíóúñ]+){1,3}),?\s+con\s+DNI", text, re.IGNORECASE)
    parts = explicit.split()
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])


def extract_loc(text: str) -> tuple[str, str]:
    patterns = [
        r"desde\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)\s+a\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)(?:\.|,|\n| el | para )",
        r"ir\s+al\s+congreso\s+de\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)\s+desde\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)(?:\s+el|\.|,|\n)",
    ]
    for i, pattern in enumerate(patterns):
        match = re.search(pattern, text, re.IGNORECASE)
        if match and i == 0:
            return match.group(1).strip(), match.group(2).strip()
        if match:
            return match.group(2).strip(), match.group(1).strip()
    return "", ""


def extract_request(path: Path) -> TravelRequest:
    text = read_email(path)
    request = TravelRequest(email_id=path.stem)
    request.source_message_id = clean_message_id(first_match(r"^Message-ID:\s*(.+)$", text, re.IGNORECASE | re.MULTILINE))
    request.delegado_rma = first_match(r"^\s*\*?\*?([A-Z]{3,5})\*?\*?\s+solicita", text, re.IGNORECASE | re.MULTILINE)
    request.gerente = manager_from_cc(text)
    request.email = (
        first_match(r"(?:De|From):\s*.*?([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", text)
        or first_match(r"(?:mi\s+email\s+es|email\s+del\s+asistente:?)\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", text)
    )
    request.nombre = label_match("NOMBRE", text)
    request.apellidos = label_match("APELLIDOS", text)
    request.dni = normalize_dni(label_match("NIF", text) or first_match(r"\bDNI\s+([0-9]{7,8}\s*[- ]?\s*[A-Z])\b", text))
    request.email = label_match("EMAIL", text) or request.email
    request.telefono = normalize_phone(label_match("MÓVIL", text) or label_match("MOVIL", text) or first_match(r"(?:Telefono|Teléfono)(?:\s+es|:)?\s*([0-9 ]{9,})", text))
    if not request.nombre:
        request.nombre, request.apellidos = guess_name(text)
    request.hospital = label_match("HOSPITAL", text)
    request.especialidad = label_match("ESPECIALIDAD", text)
    request.residente = label_match("RESIDENTE", text)
    request.socio = label_match("SOCIO", text)
    request.selas_id = label_match("SELAS ID", text) or label_match("ONE KEY / SELAS ID", text)
    request.ponente_oyente = label_match("ASISTE COMO PONENTE U OYENTE", text)
    request.consentimiento_firmado = label_match("CONSENTIMIENTO FIRMADO", text)
    request.curso = label_match("CURSO", text)
    request.congreso = label_match("CONGRESO", text)
    request.evento = (
        label_match("NOMBRE EVENTO/CONGRESO/CURSO", text)
        or first_match(r"evento\s+de\s+([^.\n]+?)(?:\s+desde|\s+del|\s+necesito|\.|,|\n)", text)
        or first_match(r"congreso\s+de\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)(?:\s+desde|\.|,|\n)", text)
        or first_match(r"simposio\s+de\s+([A-ZÁÉÍÓÚÑ][\wáéíóúñ ]+?)(?:\s+en|\.|,|\n)", text)
    )
    request.origen = label_match("CIUDAD DE ORIGEN", text)
    if not request.origen:
        request.origen, request.destino = extract_loc(text)
    request.fecha_viaje = normalize_date(label_match("FECHA DEL SERVICIO", text)) or extract_date(text)
    request.departamento = label_match("DEPARTAMENTO", text)
    request.cost_center = label_match("COST CENTRE", text) or label_match("COST CENTER", text)
    request.cost_center_completo = label_match("COST CENTRE COMPLETO", text) or label_match("COST CENTER COMPLETO", text)
    request.internal_order = label_match("INTERNAL ORDER", text)
    request.proyecto = label_match("PROYECTO", text)
    request.restricciones_alimentarias = label_match("RESTRICCIONES ALIMENTARIAS", text)
    request.descripcion_servicio = label_match("DESCRPICIÓN DEL SERVICIO", text) or label_match("DESCRIPCIÓN DEL SERVICIO", text) or label_match("DESCRIPCION DEL SERVICIO", text)
    extract_service_details(request)
    if request.descripcion_servicio and not request.destino:
        route = re.search(r"\b([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+)\s*-\s*([A-ZÁÉÍÓÚÑ][\wáéíóúñ]+)\b", request.descripcion_servicio)
        if route:
            request.destino = route.group(2)
    request.observaciones = request.descripcion_servicio
    request.hora_preferida = first_match(r"(?:sobre las|entre las)\s+([0-9]{1,2}:[0-9]{2}(?:\s+y\s+las\s+[0-9]{1,2}:[0-9]{2})?)", text)
    request.hora_preferida = re.sub(r"\s+y\s+las\s+", "-", request.hora_preferida)
    request.preferencia = "tren directo" if re.search(r"tren.{0,40}direct[oa]", text, re.IGNORECASE) else ""
    if not request.observaciones:
        request.observaciones = "Extraido por reglas locales."
    return validate_request(request)


def request_from_dict(email_id: str, data: dict[str, object]) -> TravelRequest:
    clean = {field: str(data.get(field, "") or "").strip() for field in OUTPUT_FIELDS if field != "email_id"}
    clean["dni"] = normalize_dni(clean["dni"])
    clean["telefono"] = normalize_phone(clean["telefono"])
    clean["fecha_viaje"] = normalize_date(clean["fecha_viaje"], default_year=None)
    clean["source_message_id"] = clean_message_id(clean.get("source_message_id", ""))
    return TravelRequest(email_id=email_id, **clean)


def validate_request(request: TravelRequest) -> TravelRequest:
    missing = [field for field in MANDATORY_FIELDS if not getattr(request, field).strip()]
    doubts = [request.dudas.strip()] if request.dudas.strip() else []

    invalid_dni = bool(request.dni) and not re.fullmatch(r"[0-9]{7,8}[A-Z]", request.dni)
    if invalid_dni:
        doubts.append("DNI con formato dudoso")

    invalid_date = bool(request.fecha_viaje) and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", request.fecha_viaje)
    if invalid_date:
        doubts.append("fecha_viaje no normalizada a YYYY-MM-DD")

    doubts.extend(f"falta {field}" for field in missing)
    request.estado = "pendiente_revision" if missing or invalid_dni or invalid_date else "ok"
    request.dudas = "; ".join(dict.fromkeys(doubts))
    return request


def parse_openai_text(response: dict[str, object]) -> str:
    if isinstance(response.get("output_text"), str):
        return str(response["output_text"])

    output = response.get("output", [])
    if not isinstance(output, list):
        return ""

    parts: list[str] = []
    for item in output:
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []):
            if isinstance(content, dict) and content.get("type") in {"output_text", "text"}:
                text = content.get("text", "")
                if isinstance(text, str):
                    parts.append(text)
    return "\n".join(parts).strip()


def extract_with_openai(path: Path, api_key: str) -> TravelRequest:
    email_text = read_email(path)
    payload = {
        "model": OPENAI_MODEL,
        "input": [
            {"role": "system", "content": EXTRACTION_PROMPT},
            {"role": "user", "content": f"Email a procesar:\n\n{email_text}"},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": OPENAI_JSON_SCHEMA["name"],
                "schema": OPENAI_JSON_SCHEMA["schema"],
                "strict": True,
            }
        },
    }
    request = urllib.request.Request(
        OPENAI_API_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    with urllib.request.urlopen(request, timeout=OPENAI_TIMEOUT_SECONDS) as response:
        body = json.loads(response.read().decode("utf-8"))

    extracted_text = parse_openai_text(body)
    extracted_data = json.loads(extracted_text)
    travel_request = request_from_dict(path.stem, extracted_data)
    note = "Extraido con OpenAI API."
    travel_request.observaciones = f"{travel_request.observaciones} {note}".strip() if travel_request.observaciones else note
    return validate_request(travel_request)


def extract_request_auto(path: Path, use_openai: bool) -> TravelRequest:
    text = read_email(path)
    if "DATOS PARA PETICIONES DE INVITADOS" in text.upper():
        return extract_request(path)

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if use_openai and api_key:
        try:
            return extract_with_openai(path, api_key)
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, KeyError, ValueError) as exc:
            fallback = extract_request(path)
            fallback.observaciones = f"{fallback.observaciones} Fallback local por error OpenAI: {exc}"
            return validate_request(fallback)

    return extract_request(path)


def nn_template_path() -> Path | None:
    configured = os.getenv("NN_TEMPLATE_PATH", "").strip()
    candidates = [Path(configured)] if configured else []
    candidates.append(DEFAULT_NN_TEMPLATE_PATH)
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = sheet.cell(row=source_row, column=col)
        target = sheet.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)


def apply_nn_font_style(sheet, max_col: int, max_row: int) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            color = copy(cell.font.color) if cell.font and cell.font.color else None
            cell.font = Font(name="Calibri", size=10, bold=False, color=color, underline=None)

    for row in range(1, 3):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            color = copy(cell.font.color) if cell.font and cell.font.color else None
            cell.font = Font(name="Calibri", size=11, bold=True, color=color, underline=None)
    sheet["A2"].font = Font(name="Calibri", size=22, bold=True, color="000000", underline=None)


def apply_nn_row_banding(sheet, max_col: int, first_row: int, last_row: int) -> None:
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    soft_gray_fill = PatternFill("solid", fgColor="F6F7F9")
    for row_num in range(first_row, last_row + 1):
        fill = white_fill if (row_num - first_row) % 2 == 0 else soft_gray_fill
        sheet.row_dimensions[row_num].height = 44
        for col in range(1, max_col + 1):
            sheet.cell(row=row_num, column=col).fill = fill


def event_output_root() -> Path:
    configured = os.getenv("EVENT_OUTPUT_DIR", "").strip()
    return Path(configured) if configured else EVENT_OUTPUT_DIR


def processed_ids_path() -> Path:
    configured = os.getenv("PROCESSED_IDS_PATH", "").strip()
    if configured:
        return Path(configured)
    configured_event_root = os.getenv("EVENT_OUTPUT_DIR", "").strip()
    root = Path(configured_event_root) if configured_event_root else OUTPUT_DIR
    return root / PROCESSED_IDS_FILENAME


def load_processed_message_ids() -> set[str]:
    path = processed_ids_path()
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return set()
    if isinstance(data, list):
        return {clean_message_id(item) for item in data if clean_message_id(item)}
    if isinstance(data, dict):
        return {clean_message_id(item) for item in data.get("message_ids", []) if clean_message_id(item)}
    return set()


def save_processed_message_ids(message_ids: set[str]) -> None:
    path = processed_ids_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "message_ids": sorted(message_ids),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def mark_processed_message_ids(rows: list[TravelRequest]) -> None:
    message_ids = load_processed_message_ids()
    for row in rows:
        message_id = clean_message_id(row.source_message_id)
        if message_id:
            message_ids.add(message_id)
    save_processed_message_ids(message_ids)


def clear_local_event_outputs(event_root: Path) -> None:
    if event_root.resolve() != EVENT_OUTPUT_DIR.resolve() or not event_root.exists():
        return
    for path in event_root.glob("*.xlsx"):
        path.unlink()


def apply_bot_added_font(sheet, row_num: int, max_col: int, has_dietary_restriction: bool) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=row_num, column=col)
        cell.font = Font(name="Calibri", size=10, color=BOT_ADDED_FONT_COLOR, bold=False, underline=None)
    sheet[f"AO{row_num}"].font = Font(
        name="Calibri",
        size=10,
        color=BOT_ADDED_FONT_COLOR,
        bold=has_dietary_restriction,
        underline=None,
    )


def write_nn_excel(rows: list[TravelRequest], output_path: Path = XLSX_PATH, title: str | None = None) -> bool:
    template = nn_template_path()
    if not template:
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template)
    sheet = workbook["Totales"] if "Totales" in workbook.sheetnames else workbook.active
    sheet.title = "Totales"
    sheet.conditional_formatting._cf_rules.clear()
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = "A3:DP3"
    sheet["A2"] = excel_text(title or "LISTADO GLOBAL")

    max_col = sheet.max_column
    last_display_row = max(40, 3 + len(rows))
    for row_num in range(4, last_display_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row_num, column=col).value = None

    apply_nn_row_banding(sheet, max_col, 4, last_display_row)

    for index, item in enumerate(rows, start=4):
        copy_row_style(sheet, 4, index, max_col)
        values = {
            "A": index - 3,
            "B": billing_month(item.fecha_viaje),
            "D": excel_text(item.evento),
            "E": item.delegado_rma,
            "F": excel_text(item.gerente),
            "G": item.cost_center,
            "H": item.internal_order,
            "I": item.selas_id,
            "J": excel_text(item.ponente_oyente),
            "L": excel_text(item.especialidad),
            "M": excel_text(item.nombre),
            "N": excel_text(item.apellidos),
            "O": item.dni,
            "U": excel_text(item.email, email=True),
            "V": item.telefono,
            "W": excel_text(item.origen),
            "X": parse_iso_date(item.fecha_viaje),
            "AA": excel_text(item.conex_ida),
            "AB": excel_text(item.desplazamientos_ida),
            "AC": excel_text(item.desplazamientos_vuelta),
            "AD": excel_text(item.conex_regreso),
            "AE": excel_text(item.hotel),
            "AF": parse_iso_date(item.hotel_in),
            "AG": parse_iso_date(item.hotel_out),
            "AH": excel_text(item.inscripcion),
            "AI": excel_text(item.socio),
            "AO": excel_text(item.restricciones_alimentarias),
            "AR": excel_text(item.observaciones),
        }
        for column, value in values.items():
            sheet[f"{column}{index}"] = value

        sheet[f"B{index}"].number_format = EXCEL_MONTH_FORMAT
        for date_column in ("X", "AF", "AG"):
            sheet[f"{date_column}{index}"].number_format = EXCEL_DATE_FORMAT
        has_dietary_restriction = item.restricciones_alimentarias.strip().lower() not in {"", "no", "ninguna", "ninguno"}
        sheet[f"AR{index}"].alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 7,
        "B": 18,
        "D": 22,
        "E": 16,
        "G": 15,
        "H": 17,
        "I": 18,
        "J": 18,
        "L": 24,
        "M": 18,
        "N": 26,
        "O": 14,
        "U": 28,
        "V": 14,
        "W": 18,
        "X": 16,
        "AA": 14,
        "AB": 44,
        "AC": 44,
        "AD": 16,
        "AE": 34,
        "AF": 14,
        "AG": 14,
        "AH": 16,
        "AO": 18,
        "AR": 70,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    apply_nn_row_banding(sheet, max_col, 4, last_display_row)

    apply_nn_font_style(sheet, max_col, last_display_row)
    for index, item in enumerate(rows, start=4):
        has_dietary_restriction = item.restricciones_alimentarias.strip().lower() not in {"", "no", "ninguna", "ninguno"}
        apply_bot_added_font(sheet, index, max_col, has_dietary_restriction)
        sheet[f"AR{index}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet[f"B{index}"].number_format = EXCEL_MONTH_FORMAT
        for date_column in ("X", "AF", "AG"):
            sheet[f"{date_column}{index}"].number_format = EXCEL_DATE_FORMAT

    workbook.save(output_path)
    return True


def write_outputs(rows: list[TravelRequest]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = [{field: asdict(row).get(field, "") for field in OUTPUT_FIELDS} for row in rows]

    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    headers = OUTPUT_FIELDS
    with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

    if write_nn_excel(rows, XLSX_PATH, "LISTADO GLOBAL"):
        by_event: dict[str, list[TravelRequest]] = {}
        for row in rows:
            event_name = row.evento.strip() or "SIN EVENTO"
            by_event.setdefault(event_name, []).append(row)

        event_root = event_output_root()
        clear_local_event_outputs(event_root)
        for event_name, event_rows in by_event.items():
            event_path = event_root / event_workbook_filename(event_name, event_rows)
            write_nn_excel(event_rows, event_path, event_name)
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Viajes"
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for item in data:
        sheet.append([item[header] for header in headers])

    status_col = headers.index("estado") + 1
    for row in range(2, sheet.max_row + 1):
        status = sheet.cell(row=row, column=status_col).value
        fill = PatternFill("solid", fgColor="C6EFCE" if status == "ok" else "FFEB9C")
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = fill

    for col in range(1, sheet.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(sheet.cell(row=row, column=col).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

    sheet.freeze_panes = "A2"
    workbook.save(XLSX_PATH)


def next_email_id() -> int:
    existing = []
    for path in EMAIL_DIR.glob("email_*.txt"):
        match = re.search(r"email_(\d+)", path.stem)
        if match:
            existing.append(int(match.group(1)))
    return max(existing, default=0) + 1


def make_demo_dni(index: int) -> str:
    letters = "TRWAGMYFPDXBNJZSQVHLCKE"
    number = 20000000 + index * 137
    return f"{number}{letters[number % 23]}"


def generate_demo_email(index: int, email_number: int) -> str:
    first = DEMO_FIRST_NAMES[index % len(DEMO_FIRST_NAMES)]
    last = DEMO_LAST_NAMES[index % len(DEMO_LAST_NAMES)]
    origin = DEMO_CITIES[index % len(DEMO_CITIES)]
    destination = "Rio"
    event = DEMO_CONGRESSES[index % len(DEMO_CONGRESSES)]
    hospital = DEMO_HOSPITALS[index % len(DEMO_HOSPITALS)]
    specialty = DEMO_EVENTS[index % len(DEMO_EVENTS)]
    day = 10 + (index % 15)
    phone = 600000000 + email_number
    dni = make_demo_dni(email_number)
    attendee_email = f"{first.lower()}.{last.split()[0].lower()}@example.com"
    delegate = ["EANZ", "OAJS", "MPRT", "LCSA"][index % 4]
    selas_id = f"WESM{30000000 + email_number}"
    cost_center = f"012-{25 + (index % 8):04d}"
    internal_order = f"012-S2026{80 + index:03d}"
    dietary = "No" if index % 4 else "Sin gluten"
    role = "Ponente" if index % 3 == 0 else "Oyente"

    if index % 5 == 4:
        return f"""De: Alejandra Atienza Garcia-Cuadrado {delegate}@novonordisk.com
Enviado el: jueves, 2 de julio de 2026 14:11
Para: operaciones@micetravel.es
CC: EANZ@novonordisk.com; OAJS@novonordisk.com
Asunto: NUEVA SOLICITUD: {event} - {first.upper()} {last.upper()}

image002.png

Estimado compañero/a

{delegate} solicita el siguiente evento para invitados:

DATOS PARA PETICIONES DE INVITADOS

NOMBRE: {first.upper()}
APELLIDOS: {last.upper()}
NIF: {dni}
EMAIL: {attendee_email}
MÓVIL: {phone}
HOSPITAL: {hospital}
ESPECIALIDAD: {specialty}
RESIDENTE: No
SOCIO: No
SELAS ID: {selas_id}
ASISTE COMO PONENTE U OYENTE: {role}
CONSENTIMIENTO FIRMADO: Sí
CURSO: NO
CONGRESO: SI
NOMBRE EVENTO/CONGRESO/CURSO: {event}
CIUDAD DE ORIGEN: {origin}
DEPARTAMENTO: Customer Strategy
COST CENTRE: {cost_center}
COST CENTRE COMPLETO: {cost_center} Marketing Obesity
INTERNAL ORDER: {internal_order}
PROYECTO:
RESTRICCIONES ALIMENTARIAS: {dietary}
DESCRPICIÓN DEL SERVICIO: Vuelo IDA con salida desde {origin} a {destination}. Pendiente confirmar fecha exacta del servicio.
"""

    return f"""De: Alejandra Atienza Garcia-Cuadrado {delegate}@novonordisk.com
Enviado el: jueves, 2 de julio de 2026 14:11
Para: operaciones@micetravel.es
CC: EANZ@novonordisk.com; OAJS@novonordisk.com
Asunto: NUEVA SOLICITUD: {event} - {first.upper()} {last.upper()}

image002.png

Estimado compañero/a

{delegate} solicita el siguiente evento para invitados:

DATOS PARA PETICIONES DE INVITADOS

NOMBRE: {first.upper()}
APELLIDOS: {last.upper()}
NIF: {dni}
EMAIL: {attendee_email}
MÓVIL: {phone}
HOSPITAL: {hospital}
ESPECIALIDAD: {specialty}
RESIDENTE: No
SOCIO: No
SELAS ID: {selas_id}
ASISTE COMO PONENTE U OYENTE: {role}
CONSENTIMIENTO FIRMADO: Sí
CURSO: NO
CONGRESO: SI
NOMBRE EVENTO/CONGRESO/CURSO: {event}
CIUDAD DE ORIGEN: {origin}
FECHA DEL SERVICIO: {day:02d}/09/2026
DEPARTAMENTO: Customer Strategy
COST CENTRE: {cost_center}
COST CENTRE COMPLETO: {cost_center} Marketing Obesity
INTERNAL ORDER: {internal_order}
PROYECTO:
RESTRICCIONES ALIMENTARIAS: {dietary}
DESCRPICIÓN DEL SERVICIO: Alojamiento del {day + 1:02d} de septiembre al 04 de octubre en el hotel seleccionado. Vuelo IDA {day:02d} de septiembre {origin}-{destination}. Vuelo VUELTA {destination}-{origin} 04 de octubre con llegada 05 de octubre. Vuelo directo. Turista, salida de emergencia, o en su defecto, pasillo. No es necesario tramitar inscripción en el congreso.
"""


def generate_demo_emails(count: int = 10) -> list[Path]:
    EMAIL_DIR.mkdir(parents=True, exist_ok=True)
    start = next_email_id()
    created = []
    offset = random.randint(0, 99)
    for index in range(count):
        email_number = start + index
        path = EMAIL_DIR / f"email_{email_number:03d}.txt"
        path.write_text(generate_demo_email(offset + index, email_number), encoding="utf-8")
        created.append(path)
    return created


def parse_raw_demo_email(raw: str) -> dict[str, str]:
    headers, _, body = raw.partition("\n\n")
    result = {"from": "", "to": "", "subject": "Datos viaje", "body": body.strip()}
    for line in headers.splitlines():
        key, _, value = line.partition(":")
        normalized = key.strip().lower()
        if normalized == "de":
            result["from"] = value.strip()
        elif normalized == "para":
            result["to"] = value.strip()
        elif normalized == "asunto":
            result["subject"] = value.strip()
    return result


def mail_env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def log_event(message: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")


def recent_logs(limit: int = 80) -> list[str]:
    if not LOG_PATH.exists():
        return []
    return LOG_PATH.read_text(encoding="utf-8").splitlines()[-limit:]


def require_mail_config() -> None:
    require_imap_config()


def require_imap_config() -> None:
    missing = [
        name
        for name in ("IMAP_USER", "IMAP_PASSWORD")
        if not mail_env(name)
    ]
    if missing:
        raise ValueError("Faltan variables en .env: " + ", ".join(missing))


def imap_search_unseen(mailbox: imaplib.IMAP4_SSL) -> list[bytes]:
    status, data = mailbox.uid("search", None, "UNSEEN", "SUBJECT", '"NUEVA SOLICITUD"')
    if status != "OK":
        raise RuntimeError("No se pudo buscar correo en IMAP")
    return data[0].split()


def open_imap_mailbox() -> imaplib.IMAP4_SSL:
    require_imap_config()
    imap_host = mail_env("IMAP_HOST", "imap.gmail.com")
    imap_port = int(mail_env("IMAP_PORT", "993"))
    imap_user = mail_env("IMAP_USER")
    imap_password = mail_env("IMAP_PASSWORD")
    imap_folder = mail_env("IMAP_FOLDER", "INBOX")
    mailbox = imaplib.IMAP4_SSL(imap_host, imap_port)
    mailbox.login(imap_user, imap_password)
    mailbox.select(imap_folder)
    return mailbox


def check_unread_mail() -> int:
    with open_imap_mailbox() as mailbox:
        count = len(imap_search_unseen(mailbox))
    log_event(f"Detectados {count} correos no leidos con asunto NUEVA SOLICITUD")
    return count


def test_imap_connection() -> str:
    with open_imap_mailbox():
        pass
    user = mail_env("IMAP_USER")
    log_event(f"Conexion IMAP correcta para {user}")
    return user


def imap_is_ok() -> bool:
    try:
        with open_imap_mailbox():
            return True
    except Exception:
        return False


def excel_is_ready() -> bool:
    return bool(nn_template_path() or XLSX_PATH.exists())


def run_bot_once() -> tuple[int, list[Path], list[TravelRequest]]:
    try:
        unread_before = check_unread_mail()
        imported = import_new_mail()
        rows = process_all()
        log_event(f"Bot ON: {unread_before} no leidos detectados, {len(imported)} importados")
        return unread_before, imported, rows
    except Exception as exc:
        notify_bot_error("Error ejecutando lectura de correo y volcado a Excel", exc)
        raise


def bot_is_active() -> bool:
    with BOT_LOCK:
        return BOT_ACTIVE


def bot_loop() -> None:
    global BOT_ACTIVE
    log_event(f"Bot automatico iniciado. Revision cada {WATCH_INTERVAL_SECONDS} segundos")
    try:
        while not BOT_STOP_EVENT.wait(WATCH_INTERVAL_SECONDS):
            try:
                run_bot_once()
            except Exception as exc:
                log_event(f"Error en revision automatica: {exc}")
    finally:
        with BOT_LOCK:
            BOT_ACTIVE = False
        log_event("Bot automatico detenido")


def start_bot() -> tuple[int, list[Path], list[TravelRequest]]:
    global BOT_ACTIVE, BOT_THREAD
    result = run_bot_once()
    with BOT_LOCK:
        if not BOT_ACTIVE:
            BOT_STOP_EVENT.clear()
            BOT_ACTIVE = True
            BOT_THREAD = threading.Thread(target=bot_loop, daemon=True)
            BOT_THREAD.start()
    return result


def stop_bot() -> None:
    global BOT_ACTIVE
    BOT_STOP_EVENT.set()
    with BOT_LOCK:
        BOT_ACTIVE = False
    log_event("Bot OFF")


def send_alert_email(subject: str, body: str) -> bool:
    smtp_user = mail_env("SMTP_USER")
    smtp_password = mail_env("SMTP_PASSWORD")
    alert_to = mail_env("ALERT_EMAIL_TO", mail_env("MAIL_TO") or mail_env("IMAP_USER"))
    if not smtp_user or not smtp_password or not alert_to:
        log_event("No se envio alerta por email: faltan SMTP_USER, SMTP_PASSWORD o ALERT_EMAIL_TO")
        return False

    smtp_host = mail_env("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(mail_env("SMTP_PORT", "587"))
    mail_from = mail_env("ALERT_EMAIL_FROM", mail_env("MAIL_FROM", smtp_user))

    message = EmailMessage()
    message["From"] = mail_from
    message["To"] = alert_to
    message["Subject"] = subject
    message.set_content(body)

    if smtp_port == 465:
        server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
    else:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)

    with server:
        if smtp_port != 465:
            server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(message)
    log_event(f"Alerta enviada por email a {alert_to}")
    return True


def notify_bot_error(context: str, exc: Exception) -> None:
    details = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    log_event(f"{context}: {exc}")
    body = "\n".join(
        [
            "El bot de MICE Travel ha encontrado un problema.",
            "",
            f"Momento: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Contexto: {context}",
            f"Error: {exc}",
            "",
            "Detalle tecnico:",
            details,
        ]
    )
    try:
        send_alert_email(f"MICE Travel Bot - problema detectado", body)
    except Exception as alert_exc:
        log_event(f"No se pudo enviar la alerta por email: {alert_exc}")


def send_demo_emails(count: int = 10) -> int:
    missing = [
        name
        for name in ("SMTP_USER", "SMTP_PASSWORD", "MAIL_TO")
        if not mail_env(name)
    ]
    if missing:
        raise ValueError("Faltan variables SMTP en .env: " + ", ".join(missing))
    smtp_host = mail_env("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(mail_env("SMTP_PORT", "587"))
    smtp_user = mail_env("SMTP_USER")
    smtp_password = mail_env("SMTP_PASSWORD")
    mail_from = mail_env("MAIL_FROM", smtp_user)
    mail_to = mail_env("MAIL_TO")
    start = next_email_id()
    offset = random.randint(0, 99)

    if smtp_port == 465:
        server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
    else:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)

    with server:
        if smtp_port != 465:
            server.starttls()
        server.login(smtp_user, smtp_password)
        for index in range(count):
            raw = generate_demo_email(offset + index, start + index)
            parsed = parse_raw_demo_email(raw)
            message = EmailMessage()
            message["From"] = mail_from
            message["To"] = mail_to
            message["Reply-To"] = parsed["from"]
            message["Subject"] = parsed["subject"]
            message.set_content(parsed["body"])
            server.send_message(message)
    return count


def message_body(message: EmailMessage) -> str:
    if message.is_multipart():
        for part in message.walk():
            if part.get_content_type() == "text/plain" and not part.get_filename():
                return part.get_content()
        return ""
    return message.get_content()


def import_new_mail() -> list[Path]:
    created: list[Path] = []
    EMAIL_DIR.mkdir(parents=True, exist_ok=True)
    processed_ids = load_processed_message_ids()

    with open_imap_mailbox() as mailbox:
        for uid in imap_search_unseen(mailbox):
            uid_text = uid.decode("ascii")
            path = EMAIL_DIR / f"mail_{uid_text}.txt"
            if path.exists():
                mailbox.uid("store", uid, "+FLAGS", "(\\Seen)")
                continue
            status, message_data = mailbox.uid("fetch", uid, "(RFC822)")
            if status != "OK" or not message_data:
                continue
            raw_bytes = message_data[0][1]
            message = BytesParser(policy=policy.default).parsebytes(raw_bytes)
            message_id = clean_message_id(str(message.get("Message-ID") or f"imap-{mail_env('IMAP_USER')}-{uid_text}"))
            if message_id in processed_ids:
                mailbox.uid("store", uid, "+FLAGS", "(\\Seen)")
                log_event(f"Correo omitido por duplicado: {message_id}")
                continue
            from_value = str(message.get("Reply-To") or message.get("From") or "")
            to_value = str(message.get("To") or "")
            cc_value = str(message.get("Cc") or "")
            subject = str(message.get("Subject") or "")
            body = message_body(message)
            path.write_text(
                f"Message-ID: {message_id}\nDe: {from_value}\nPara: {to_value}\nCC: {cc_value}\nAsunto: {subject}\n\n{body}",
                encoding="utf-8",
            )
            created.append(path)
            mailbox.uid("store", uid, "+FLAGS", "(\\Seen)")
    log_event(f"Importados {len(created)} correos nuevos y marcados como leidos")
    return created


def send_and_process_demo_emails(count: int = 10) -> tuple[int, list[Path], list[TravelRequest]]:
    sent = send_demo_emails(count)
    imported: list[Path] = []

    for attempt in range(6):
        if attempt:
            time.sleep(3)
        imported = import_new_mail()
        if imported:
            break

    rows = process_paths(imported) if imported else process_all()
    return sent, imported, rows


def process_all() -> list[TravelRequest]:
    use_openai = bool(os.getenv("OPENAI_API_KEY", "").strip())
    paths = sorted(EMAIL_DIR.glob("*.txt"))
    return process_paths(paths, use_openai)


def process_paths(paths: list[Path], use_openai: bool | None = None) -> list[TravelRequest]:
    if use_openai is None:
        use_openai = bool(os.getenv("OPENAI_API_KEY", "").strip())
    rows = [extract_request_auto(path, use_openai) for path in paths]
    write_outputs(rows)
    mark_processed_message_ids(rows)
    print(f"Procesados: {len(rows)} emails")
    print(f"Extractor: {'OpenAI API' if use_openai else 'fallback local'}")
    print(f"Excel: {XLSX_PATH}")
    print(f"Excel eventos: {event_output_root()}")
    print(f"CSV: {CSV_PATH}")
    print(f"JSON: {JSON_PATH}")
    log_event(f"Excel global y listados por evento actualizados con {len(rows)} solicitudes")
    return rows


def open_excel_file() -> None:
    if not XLSX_PATH.exists():
        process_all()
    if sys.platform == "darwin":
        subprocess.run(["open", str(XLSX_PATH)], check=False)
    elif sys.platform.startswith("win"):
        os.startfile(str(XLSX_PATH))  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", str(XLSX_PATH)], check=False)
    log_event("Excel abierto desde el panel")


def rows_as_dicts() -> list[dict[str, str]]:
    if not JSON_PATH.exists():
        process_all()
    with JSON_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def dashboard_html() -> str:
    rows = rows_as_dicts()
    openai_ready = bool(os.getenv("OPENAI_API_KEY", "").strip())
    mail_ready = imap_is_ok()
    excel_ready = excel_is_ready()
    bot_active = bot_is_active()
    generated_at = time.strftime("%H:%M:%S")
    ok_count = sum(1 for row in rows if row.get("estado") == "ok")
    pending_count = len(rows) - ok_count
    preview_fields = [
        ("nombre", "Nombre"),
        ("apellidos", "Apellidos"),
        ("evento", "Evento"),
        ("origen", "Origen"),
        ("fecha_viaje", "Fecha"),
        ("desplazamientos_ida", "Ida"),
        ("desplazamientos_vuelta", "Vuelta"),
        ("hotel", "Hotel"),
        ("restricciones_alimentarias", "Alergias"),
        ("estado", "Estado"),
    ]
    table_rows = "\n".join(
        f"<tr>"
        + "".join(
            f"<td class=\"{field}\">{html.escape(str(row.get(field, '') or ''))}</td>"
            for field, _label in preview_fields
        )
        + "</tr>"
        for row in rows[-12:]
    )
    headers = "".join(f"<th>{html.escape(label)}</th>" for _field, label in preview_fields)
    logs = "\n".join(html.escape(line) for line in recent_logs(50)) or "Sin actividad todavia."
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>MICE Travel Bot</title>
  <style>
    :root {{
      color-scheme: dark;
      --ink: #f3f7fb;
      --muted: #94a9b8;
      --line: #233848;
      --soft-line: #1c3040;
      --brand: #00adef;
      --brand-soft: #7bdcff;
      --brand-dark: #062c3d;
      --surface: #061019;
      --panel: #0b1924;
      --panel-2: #102432;
      --row-alt: #0e2130;
      --warn: #f6c860;
      --ok: #52d28f;
      --danger: #ff6f7d;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Avenir Next", Avenir, "Segoe UI Variable", "Segoe UI", ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, Arial, Helvetica, sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 12% 0%, rgba(0,173,239,.18), transparent 34%),
        linear-gradient(180deg, #07131d 0%, var(--surface) 46%, #050b11 100%);
      -webkit-font-smoothing: antialiased;
    }}
    .shell {{
      min-height: 100vh;
      display: grid;
      grid-template-rows: auto auto 1fr;
    }}
    header {{
      padding: 26px 32px 20px;
      background: linear-gradient(180deg, rgba(11,32,47,.96) 0%, rgba(6,16,25,.92) 100%);
      border-bottom: 1px solid var(--line);
      display: flex;
      justify-content: space-between;
      gap: 20px;
      align-items: flex-end;
    }}
    .brand-logo {{
      display: block;
      width: 210px;
      max-width: 42vw;
      height: auto;
    }}
    .meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      font-size: 14px;
      justify-content: flex-end;
    }}
    .pill {{
      border: 1px solid var(--line);
      background: rgba(0,173,239,.06);
      border-radius: 999px;
      padding: 7px 11px;
      color: var(--muted);
    }}
    .health {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      border: 1px solid var(--line);
      background: rgba(0,173,239,.06);
      border-radius: 999px;
      padding: 7px 11px;
      color: var(--muted);
    }}
    .dot {{
      width: 9px;
      height: 9px;
      border-radius: 50%;
      background: var(--danger);
      box-shadow: 0 0 0 4px rgba(255,138,138,.12);
    }}
    .dot.ok {{
      background: var(--ok);
      box-shadow: 0 0 0 4px rgba(103,214,155,.12);
    }}
    .toolbar {{
      display: grid;
      grid-template-columns: 220px auto auto auto 1fr;
      gap: 14px;
      align-items: center;
      padding: 16px 32px;
      border-bottom: 1px solid var(--line);
      background: rgba(6,16,25,.94);
    }}
    .power {{
      min-height: 52px;
      border-radius: 999px;
      gap: 12px;
      font-size: 16px;
      justify-content: flex-start;
      padding: 8px 10px;
    }}
    .switch-track {{
      width: 58px;
      height: 32px;
      border-radius: 999px;
      background: #04101a;
      border: 1px solid rgba(123,220,255,.16);
      display: inline-flex;
      align-items: center;
      padding: 3px;
    }}
    .switch-knob {{
      width: 24px;
      height: 24px;
      border-radius: 50%;
      background: var(--danger);
      box-shadow: 0 0 14px rgba(255,138,138,.45);
      transition: transform .18s ease, background .18s ease;
    }}
    .power.is-on .switch-knob {{
      transform: translateX(26px);
      background: var(--ok);
      box-shadow: 0 0 16px rgba(103,214,155,.55);
    }}
    .power-text {{
      display: inline-flex;
      flex-direction: column;
      align-items: flex-start;
      line-height: 1.1;
    }}
    .power-text strong {{
      font-size: 16px;
    }}
    .power-text span {{
      margin-top: 3px;
      font-size: 12px;
      color: rgba(6,17,19,.72);
      font-weight: 650;
    }}
    button, a.button {{
      min-height: 40px;
      border: 1px solid rgba(123,220,255,.78);
      background: var(--brand);
      color: #031019;
      padding: 9px 14px;
      border-radius: 6px;
      font-size: 14px;
      font-weight: 760;
      cursor: pointer;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      white-space: nowrap;
      box-shadow: 0 10px 26px rgba(0,173,239,.18);
    }}
    button.secondary, a.secondary {{
      background: #082234;
      color: var(--brand-soft);
      border-color: #154d69;
    }}
    button.ghost {{
      border-color: var(--line);
      color: var(--ink);
      background: #0b1b29;
      box-shadow: none;
    }}
    button.health-action {{
      min-height: auto;
      border: 1px solid var(--line);
      background: rgba(0,173,239,.06);
      border-radius: 999px;
      padding: 7px 11px;
      color: var(--muted);
      box-shadow: none;
      font-size: 14px;
      font-weight: 500;
    }}
    button:hover, a.button:hover {{
      filter: brightness(1.06);
    }}
    #status {{
      color: var(--muted);
      font-size: 14px;
      text-align: right;
    }}
    main {{
      padding: 24px 32px 32px;
      display: grid;
      grid-template-columns: 1fr 340px;
      gap: 18px;
      min-width: 0;
    }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      box-shadow: 0 14px 34px rgba(0,0,0,.24);
    }}
    .metric .label {{
      color: var(--muted);
      font-size: 13px;
    }}
    .metric .value {{
      margin-top: 6px;
      font-size: 28px;
      font-weight: 700;
    }}
    .metric .ok {{ color: var(--ok); }}
    .metric .warn {{ color: var(--warn); }}
    section, aside {{
      min-width: 0;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      max-height: calc(100vh - 245px);
      box-shadow: 0 14px 34px rgba(0,0,0,.24);
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      min-width: 1180px;
      font-size: 14px;
    }}
    th, td {{
      border-bottom: 1px solid var(--soft-line);
      padding: 11px 12px;
      text-align: left;
      vertical-align: top;
      white-space: normal;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #0d2b3d;
      color: #d9f5ff;
      z-index: 1;
      font-size: 13px;
      white-space: nowrap;
      font-weight: 760;
    }}
    tbody tr:nth-child(even) {{ background: var(--row-alt); }}
    tbody tr:hover {{ background: #12314a; }}
    td.nombre, td.apellidos {{ white-space: nowrap; }}
    td.estado {{
      font-weight: 700;
      color: var(--ok);
      white-space: nowrap;
    }}
    td.restricciones_alimentarias {{
      font-weight: 700;
    }}
    aside {{
      background: #040d14;
      color: #e9eef2;
      border-radius: 8px;
      overflow: hidden;
      border: 1px solid var(--line);
      box-shadow: 0 14px 34px rgba(0,0,0,.24);
    }}
    aside h2 {{
      margin: 0;
      padding: 15px 16px;
      font-size: 16px;
      background: #0a2030;
      border-bottom: 1px solid var(--line);
    }}
    pre {{
      margin: 0;
      padding: 14px 16px 18px;
      max-height: calc(100vh - 218px);
      overflow: auto;
      font-family: "SF Mono", Menlo, Consolas, monospace;
      font-size: 12px;
      line-height: 1.45;
      white-space: pre-wrap;
      color: #b6cede;
    }}
    @media (max-width: 980px) {{
      header {{ align-items: flex-start; flex-direction: column; }}
      .toolbar {{ grid-template-columns: 1fr; }}
      #status {{ text-align: left; }}
      main {{ grid-template-columns: 1fr; }}
      .cards {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <header>
      <div>
        <img class="brand-logo" src="/assets/logo-micetravel-blanco.png" alt="MICE Travel">
      </div>
    </header>
    <div class="toolbar">
      <button id="powerButton" class="power {'is-on' if bot_active else ''}" onclick="toggleBot()">
        <span class="switch-track"><span class="switch-knob"></span></span>
        <span class="power-text"><strong>{"ON" if bot_active else "OFF"}</strong><span>{"Cada 5 minutos" if bot_active else "Activar bot"}</span></span>
      </button>
      <span class="health"><span class="dot {'ok' if mail_ready else ''}"></span>IMAP</span>
      <span class="health"><span class="dot {'ok' if excel_ready else ''}"></span>Excel</span>
      <button class="health-action" onclick="openExcel()">Abrir Excel</button>
      <span id="status">Listo.</span>
    </div>
    <main>
      <section>
        <div class="cards">
          <div class="metric"><div class="label">Solicitudes</div><div class="value">{len(rows)}</div></div>
          <div class="metric"><div class="label">Correctas</div><div class="value ok">{ok_count}</div></div>
          <div class="metric"><div class="label">Pendientes</div><div class="value warn">{pending_count}</div></div>
        </div>
        <div class="table-wrap">
          <table aria-label="Solicitudes procesadas">
            <thead><tr>{headers}</tr></thead>
            <tbody>{table_rows}</tbody>
          </table>
        </div>
      </section>
      <aside>
        <h2>Actividad</h2>
        <pre>{logs}</pre>
      </aside>
    </main>
  </div>
  <script>
    async function callApi(path, message, reload = true) {{
      const status = document.getElementById('status');
      status.textContent = message;
      const response = await fetch(path, {{ method: 'POST' }});
      if (!response.ok) {{
        const data = await response.json().catch(() => ({{}}));
        status.textContent = data.error || 'Ha ocurrido un error.';
        return;
      }}
      if (reload) window.location.reload();
    }}
    async function toggleBot() {{
      const button = document.getElementById('powerButton');
      const isOn = button.classList.contains('is-on');
      if (isOn) {{
        button.classList.remove('is-on');
        button.querySelector('strong').textContent = 'OFF';
        button.querySelector('.power-text span').textContent = 'Deteniendo';
        await callApi('/api/stop-bot', 'Bot detenido.');
        return;
      }}
      button.classList.add('is-on');
      button.querySelector('strong').textContent = 'ON';
      button.querySelector('.power-text span').textContent = 'Procesando';
      await callApi('/api/start-bot', 'Bot activo: buscando correos, actualizando Excel y marcando como leídos...');
    }}
    function openExcel() {{
      callApi('/api/open-excel', 'Abriendo Excel...');
    }}
  </script>
</body>
</html>"""


class DashboardHandler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args: object) -> None:
        return

    def send_bytes(self, content: bytes, content_type: str, status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_json(self, payload: dict[str, object], status: int = 200) -> None:
        self.send_bytes(json.dumps(payload, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", status)

    def do_GET(self) -> None:
        if self.path == "/":
            self.send_bytes(dashboard_html().encode("utf-8"), "text/html; charset=utf-8")
            return
        if self.path == "/download/xlsx":
            self.send_bytes(XLSX_PATH.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return
        if self.path == "/download/json":
            self.send_bytes(JSON_PATH.read_bytes(), "application/json; charset=utf-8")
            return
        if self.path == "/assets/logo-micetravel-blanco.png":
            self.send_bytes((ASSETS_DIR / "logo-micetravel-blanco.png").read_bytes(), "image/png")
            return
        self.send_json({"error": "not_found"}, status=404)

    def do_POST(self) -> None:
        try:
            self.handle_post()
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def handle_post(self) -> None:
        if self.path == "/api/generate":
            created = generate_demo_emails(10)
            rows = process_all()
            self.send_json({"created": [path.name for path in created], "rows": len(rows)})
            return
        if self.path == "/api/send-and-process":
            sent, imported, rows = send_and_process_demo_emails(10)
            self.send_json({"sent": sent, "imported": [path.name for path in imported], "rows": len(rows)})
            return
        if self.path == "/api/send-demo":
            sent = send_demo_emails(10)
            self.send_json({"sent": sent})
            return
        if self.path == "/api/test-mail":
            user = test_imap_connection()
            self.send_json({"ok": True, "user": user})
            return
        if self.path == "/api/check-mail":
            count = check_unread_mail()
            self.send_json({"unread": count})
            return
        if self.path == "/api/read-mail":
            imported = import_new_mail()
            self.send_json({"imported": [path.name for path in imported], "imported_count": len(imported)})
            return
        if self.path == "/api/process-mail":
            imported = import_new_mail()
            rows = process_all()
            self.send_json({"imported": [path.name for path in imported], "imported_count": len(imported), "rows": len(rows)})
            return
        if self.path == "/api/start-bot":
            unread_before, imported, rows = start_bot()
            self.send_json({
                "on": True,
                "unread_before": unread_before,
                "imported": [path.name for path in imported],
                "imported_count": len(imported),
                "rows": len(rows),
            })
            return
        if self.path == "/api/stop-bot":
            stop_bot()
            self.send_json({"on": False})
            return
        if self.path == "/api/process":
            rows = process_all()
            self.send_json({"rows": len(rows)})
            return
        if self.path == "/api/open-excel":
            open_excel_file()
            self.send_json({"opened": str(XLSX_PATH)})
            return
        self.send_json({"error": "not_found"}, status=404)


def find_available_port(start_port: int) -> int:
    for port in range(start_port, start_port + 20):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            try:
                sock.bind((SERVER_HOST, port))
            except OSError:
                continue
            return port
    raise OSError(f"No hay puertos libres entre {start_port} y {start_port + 19}")


def serve_dashboard() -> None:
    process_all()
    port = find_available_port(int(mail_env("EMAIL_AGENT_PORT", str(DEFAULT_SERVER_PORT))))
    server = ThreadingHTTPServer((SERVER_HOST, port), DashboardHandler)
    print(f"Panel demo: http://{SERVER_HOST}:{port}")
    print("Pulsa Ctrl+C para parar.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nPanel detenido.")
    finally:
        server.server_close()


def email_snapshot() -> dict[str, float]:
    EMAIL_DIR.mkdir(parents=True, exist_ok=True)
    return {str(path): path.stat().st_mtime for path in sorted(EMAIL_DIR.glob("*.txt"))}


def watch_emails() -> None:
    print(f"Escuchando nuevos emails en: {EMAIL_DIR}")
    print("Para la demo: guarda un nuevo .txt en esa carpeta y el Excel se actualizara solo.")
    print("Pulsa Ctrl+C para parar.\n")

    previous = email_snapshot()
    process_all()

    try:
        while True:
            time.sleep(WATCH_INTERVAL_SECONDS)
            current = email_snapshot()
            if current != previous:
                print("\nCambio detectado en emails/. Regenerando salidas...")
                previous = current
                process_all()
    except KeyboardInterrupt:
        print("\nEscucha detenida.")


def main() -> None:
    load_env_file()

    if "--dashboard" in sys.argv or "--serve" in sys.argv:
        serve_dashboard()
        return

    if "--watch" in sys.argv:
        watch_emails()
        return

    process_all()


if __name__ == "__main__":
    main()
